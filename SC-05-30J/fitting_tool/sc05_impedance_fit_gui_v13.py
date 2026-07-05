# -*- coding: utf-8 -*-
"""
SC-05-30J / コモンモードチョーク測定 CSV 解析GUI v13

機能:
  1. 測定データ CSV を選択
  2. Zopen CSV を選択
  3. 位相補償 td[ns] を入力
  4. Open補正 + τ補正 + 等価回路フィット
  5. 測定CSVと同じフォルダに Excel(.xlsx) を出力

依存ライブラリ:
  pip install numpy scipy openpyxl matplotlib pillow

GUIは標準ライブラリ tkinter を使用します。
"""

from __future__ import annotations

import csv
import math
import os
import re
import sys
import tempfile
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Tuple

import numpy as np

try:
    from scipy.optimize import least_squares
except Exception as exc:  # pragma: no cover
    least_squares = None
    _SCIPY_IMPORT_ERROR = exc
else:
    _SCIPY_IMPORT_ERROR = None

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception as exc:  # pragma: no cover
    tk = None
    filedialog = None
    messagebox = None
    _TK_IMPORT_ERROR = exc
else:
    _TK_IMPORT_ERROR = None

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter, NullFormatter
except Exception as exc:  # pragma: no cover
    plt = None
    FuncFormatter = None
    NullFormatter = None
    _MATPLOTLIB_IMPORT_ERROR = exc
else:
    _MATPLOTLIB_IMPORT_ERROR = None


# -----------------------------------------------------------------------------
# 設定値
# -----------------------------------------------------------------------------
RREF_OHM_DEFAULT = 49.8       # 2026-07-01版Excelに合わせた直列基準抵抗
TD_NS_DEFAULT = 2.6           # 今回の測定で使用した目安値

# Rs制約のデフォルト値。全周波数Zshort補正時にRsが非物理的な下限へ張り付くのを防ぐ。
RS_MODE_DEFAULT = "lower_bound"  # "free", "lower_bound", "fixed"
RS_MIN_OHM_DEFAULT = 0.05       # Rs下限モードの初期値 [Ω]
RS_FIXED_OHM_DEFAULT = 0.10     # Rs固定モードの初期値 [Ω]


@dataclass
class CsvData:
    path: Path
    frequency_hz: np.ndarray
    mag_db: np.ndarray
    phase_deg: np.ndarray
    header_row: int
    freq_col: int
    mag_col: int
    phase_col: int


@dataclass
class FitResult:
    params: dict
    rms_mag_error_db: float
    rms_phase_error_deg: float
    f_main_peak_hz: float
    f_hf_min_hz: float
    rs_mode: str
    rs_min_ohm: float
    rs_fixed_ohm: float
    rs_initial_ohm: float
    rs_lower_bound_ohm: float
    rs_note: str


# -----------------------------------------------------------------------------
# 数値ユーティリティ
# -----------------------------------------------------------------------------
def wrap_deg(angle_deg: np.ndarray | float) -> np.ndarray | float:
    """角度を -180〜+180 deg に折り返す。"""
    return (np.asarray(angle_deg) + 180.0) % 360.0 - 180.0


def complex_from_mag_phase(mag_db: np.ndarray, phase_deg: np.ndarray) -> np.ndarray:
    mag = 10.0 ** (mag_db / 20.0)
    return mag * np.exp(1j * np.deg2rad(phase_deg))


def h_to_impedance(H: np.ndarray, rref_ohm: float = RREF_OHM_DEFAULT) -> np.ndarray:
    """
    測定された伝達関数 H からDUTインピーダンスへ変換する。

    想定回路:
        信号源 -- Rref -- DUT -- GND
        H = Vdut / Vin

    このとき、Zdut = Rref * H / (1 - H)
    """
    eps = 1e-30
    return rref_ohm * H / (1.0 - H + eps)


def interp_complex_logf(f_src: np.ndarray, z_src: np.ndarray, f_dst: np.ndarray) -> np.ndarray:
    """周波数軸が少し異なる場合に、複素データを log10(f) 軸で補間する。"""
    x_src = np.log10(f_src)
    x_dst = np.log10(f_dst)
    re = np.interp(x_dst, x_src, np.real(z_src))
    im = np.interp(x_dst, x_src, np.imag(z_src))
    return re + 1j * im


def model_impedance(freq_hz: np.ndarray, p: dict) -> np.ndarray:
    """
    今回使用した等価回路モデル。

        Y = 1/(Rs + s Lcm) + 1/Rp + s Cp + 1/(R2 + s L2 + 1/(s C2))
        Z = 1/Y

    モデル位相には τ補正を加えない。
    """
    w = 2.0 * np.pi * freq_hz
    s = 1j * w

    Rs = p["Rs"]
    Lcm = p["Lcm"]
    Cp = p["Cp"]
    Rp = p["Rp"]
    R2 = p["R2"]
    L2 = p["L2"]
    C2 = p["C2"]

    Y1 = 1.0 / (Rs + s * Lcm)
    Yrp = 1.0 / Rp
    Ycp = s * Cp
    Z2 = R2 + s * L2 + 1.0 / (s * C2)
    Y2 = 1.0 / Z2
    Y = Y1 + Yrp + Ycp + Y2
    return 1.0 / Y


def decompose_model(freq_hz: np.ndarray, p: dict) -> dict:
    """Excel出力用にモデル各枝のアドミタンスを分解して返す。"""
    w = 2.0 * np.pi * freq_hz
    s = 1j * w
    Y1 = 1.0 / (p["Rs"] + s * p["Lcm"])
    Yrp = np.ones_like(freq_hz, dtype=complex) / p["Rp"]
    Ycp = s * p["Cp"]
    Z2 = p["R2"] + s * p["L2"] + 1.0 / (s * p["C2"])
    Y2 = 1.0 / Z2
    Ytotal = Y1 + Yrp + Ycp + Y2
    Ztotal = 1.0 / Ytotal
    return {
        "Y1": Y1,
        "Yrp": Yrp,
        "Ycp": Ycp,
        "Z2": Z2,
        "Y2": Y2,
        "Ytotal": Ytotal,
        "Ztotal": Ztotal,
    }


# -----------------------------------------------------------------------------
# CSV読込
# -----------------------------------------------------------------------------
def _read_text_with_fallback(path: Path) -> str:
    encodings = ["utf-8-sig", "cp932", "shift_jis", "utf-8"]
    last_error = None
    for enc in encodings:
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise UnicodeDecodeError("csv", b"", 0, 1, f"CSVの文字コードを判定できません: {last_error}")


def _to_float(s: object) -> float | None:
    if s is None:
        return None
    t = str(s).strip().replace("\ufeff", "")
    if not t:
        return None
    t = t.replace(",", "")
    # 単位付き文字列への保険
    t = re.sub(r"[^0-9eE+\-.]", "", t)
    if t in ("", "+", "-", "."):
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _find_header_and_rows(rows: List[List[str]]) -> Tuple[int, List[str], List[List[str]]]:
    """Frequencyを含むヘッダ行を探す。見つからない場合は先頭の数値3列行の直前を仮ヘッダにする。"""
    for i, row in enumerate(rows):
        joined = " ".join(str(c).lower() for c in row)
        if any(k in joined for k in ["frequency", "freq", "周波数", "hz"]):
            return i, row, rows[i + 1 :]

    # ヘッダ無しCSVへの保険: 3列以上が数値になる最初の行をデータ開始とする
    for i, row in enumerate(rows):
        nums = [_to_float(c) for c in row]
        if sum(v is not None for v in nums) >= 3:
            # ヘッダ無しCSVでは列名からMag/Phaseを決め打ちしない。
            # 4列以上CSVでは 2列目がChannel 1側振幅のことがあり、
            # ここで仮に Mag/Phase と名付けると誤検出するため。
            header = [f"Col{j+1}" for j in range(len(row))]
            return max(0, i - 1), header, rows[i:]

    raise ValueError("CSV内に周波数・振幅・位相の数値データを見つけられません。")


def _column_score(name: str, kind: str) -> int:
    n = name.lower().replace(" ", "_")
    score = 0
    if kind == "freq":
        if "frequency" in n or "freq" in n or "周波数" in n:
            score += 100
        if "hz" in n:
            score += 20
    elif kind == "mag":
        if "phase" in n or "位相" in n:
            score -= 100
        if "mag" in n or "magnitude" in n or "振幅" in n or "gain" in n:
            score += 80
        if "db" in n or "dB" in name:
            score += 40
        if "ch2" in n or "channel_2" in n or "channel2" in n:
            score += 10
    elif kind == "phase":
        if "phase" in n or "位相" in n:
            score += 100
        if "deg" in n or "degree" in n or "°" in name:
            score += 20
        if "ch2" in n or "channel_2" in n or "channel2" in n:
            score += 10
    return score


def _select_columns(header: List[str], data_rows: List[List[str]]) -> Tuple[int, int, int]:
    """
    Frequency, Channel 2 Magnitude[dB], Channel 2 Phase[deg] を選ぶ。

    WaveForms/AD2のCSVには次のような複数形式があるため、ヘッダ名で判定できない場合も
    「先頭のChannel 1/入力側振幅らしい0dB近傍の列」を避け、DUT側のMagnitude/Phase列を選ぶ。

    典型例:
      3列CSV: Frequency, Channel 2 Magnitude, Channel 2 Phase
      4列以上CSV: Frequency, Channel 1 Magnitude, Channel 2 Magnitude, Channel 2 Phase, ...

    7/1版Excelと合わせるため、4列以上でヘッダ判定できない場合は
    [freq, ch1_mag, ch2_mag, ch2_phase] の並びを優先する。
    """
    if not data_rows:
        raise ValueError("CSVにデータ行がありません。")

    max_cols = max(len(r) for r in data_rows[:200])

    def col_values(c: int, nmax: int = 300) -> np.ndarray:
        vals = []
        for row in data_rows[:nmax]:
            if c < len(row):
                v = _to_float(row[c])
                if v is not None and np.isfinite(v):
                    vals.append(v)
        return np.asarray(vals, dtype=float)

    # まずヘッダ名で判定
    header_l = [str(h).lower().replace(" ", "_") for h in header]
    scores = {
        "freq": [_column_score(h, "freq") for h in header],
        "mag": [_column_score(h, "mag") for h in header],
        "phase": [_column_score(h, "phase") for h in header],
    }

    # Channel 1 と Channel 2 が同時にあるCSVでは、MagnitudeはChannel 2を強く優先する。
    for i, h in enumerate(header_l):
        is_ch2 = ("channel_2" in h) or ("ch2" in h) or ("channel2" in h)
        is_ch1 = ("channel_1" in h) or ("ch1" in h) or ("channel1" in h)
        if is_ch2:
            scores["mag"][i] += 80
            scores["phase"][i] += 80
        if is_ch1:
            scores["mag"][i] -= 80
            scores["phase"][i] -= 80

    freq_col = int(np.argmax(scores["freq"])) if max(scores["freq"]) > 0 else -1
    mag_col = int(np.argmax(scores["mag"])) if max(scores["mag"]) > 0 else -1
    phase_col = int(np.argmax(scores["phase"])) if max(scores["phase"]) > 0 else -1

    if freq_col >= 0 and mag_col >= 0 and phase_col >= 0 and len({freq_col, mag_col, phase_col}) == 3:
        return freq_col, mag_col, phase_col

    # ヘッダで分からない場合: 数値列から推定
    numeric_cols = []
    for c in range(max_cols):
        vals = col_values(c)
        if len(vals) >= 5:
            numeric_cols.append(c)

    if len(numeric_cols) < 3:
        raise ValueError("数値列が3列未満です。Frequency, Magnitude[dB], Phase[deg] が必要です。")

    # 周波数列は単調増加し、正値で、範囲が広い列を優先
    best_freq = numeric_cols[0]
    best_score = -1e99
    for c in numeric_cols:
        arr = col_values(c, 500)
        if len(arr) < 5 or np.any(arr <= 0):
            continue
        monotonic = np.mean(np.diff(arr) > 0)
        span = np.log10(np.nanmax(arr) / np.nanmin(arr)) if np.nanmin(arr) > 0 else 0
        sc = 100 * monotonic + 10 * span
        if sc > best_score:
            best_score = sc
            best_freq = c

    remaining = [c for c in numeric_cols if c != best_freq]
    if len(remaining) < 2:
        raise ValueError("振幅列と位相列を推定できません。")

    def is_flat_0db(c: int) -> bool:
        arr = col_values(c, 500)
        if len(arr) < 10:
            return False
        med = float(np.nanmedian(arr))
        p05, p95 = np.nanpercentile(arr, [5, 95])
        return abs(med) < 3.0 and (p95 - p05) < 3.0

    def phase_score(c: int) -> float:
        arr = col_values(c, 500)
        if len(arr) < 10:
            return -1e9
        p05, p95 = np.nanpercentile(arr, [5, 95])
        span = float(p95 - p05)
        med_abs = float(abs(np.nanmedian(arr)))
        # 位相は通常 -360〜+360deg付近、かつある程度変化または大きな絶対値を持つ。
        score = 0.0
        score += 20.0 if np.nanmin(arr) >= -400 and np.nanmax(arr) <= 400 else -50.0
        score += min(span, 180.0) / 3.0
        score += min(med_abs, 180.0) / 6.0
        if is_flat_0db(c):
            score -= 60.0
        return score

    def mag_score(c: int) -> float:
        arr = col_values(c, 500)
        if len(arr) < 10:
            return -1e9
        p05, p95 = np.nanpercentile(arr, [5, 95])
        span = float(p95 - p05)
        score = 0.0
        # dB振幅は概ね -120〜+20dBに入ることが多い。
        score += 20.0 if np.nanmin(arr) >= -160 and np.nanmax(arr) <= 40 else -30.0
        score += min(span, 80.0) / 2.0
        if is_flat_0db(c):
            # Channel 1 / Vin側の0dB近傍列をMagnitudeとして選ばない。
            score -= 80.0
        return score

    # 典型の4列以上CSV: Frequency, Channel1 Magnitude, Channel2 Magnitude, Channel2 Phase
    # ヘッダ不明時、最初の残り列が0dB近傍で平坦なら、それを入力側列として飛ばす。
    if len(remaining) >= 3 and is_flat_0db(remaining[0]):
        cand_mag, cand_phase = remaining[1], remaining[2]
        if phase_score(cand_phase) > phase_score(cand_mag) - 10:
            return best_freq, cand_mag, cand_phase

    # 一般スコアで最良の mag/phase ペアを選ぶ。
    best_pair = None
    best_pair_score = -1e99
    for mc in remaining:
        for pc in remaining:
            if pc == mc:
                continue
            sc = mag_score(mc) + phase_score(pc)
            # CSV上でMagnitudeの直後にPhaseが来る並びを少し優先。
            if pc == mc + 1:
                sc += 15.0
            if sc > best_pair_score:
                best_pair_score = sc
                best_pair = (mc, pc)

    if best_pair is None:
        # 最後の保険: 3列CSV想定
        return best_freq, remaining[0], remaining[1]

    return best_freq, best_pair[0], best_pair[1]

def load_ad2_csv(path_like: str | Path) -> CsvData:
    path = Path(path_like)
    text = _read_text_with_fallback(path)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except Exception:
        dialect = csv.excel
    rows = list(csv.reader(text.splitlines(), dialect))
    rows = [r for r in rows if any(str(c).strip() for c in r)]

    header_row, header, data_rows = _find_header_and_rows(rows)
    freq_col, mag_col, phase_col = _select_columns(header, data_rows)

    f_list, mag_list, phase_list = [], [], []
    for row in data_rows:
        if max(freq_col, mag_col, phase_col) >= len(row):
            continue
        f = _to_float(row[freq_col])
        m = _to_float(row[mag_col])
        ph = _to_float(row[phase_col])
        if f is None or m is None or ph is None:
            continue
        if f <= 0:
            continue
        f_list.append(f)
        mag_list.append(m)
        phase_list.append(ph)

    if len(f_list) < 10:
        raise ValueError("有効な測定点が10点未満です。CSV列の判定を確認してください。")

    f = np.asarray(f_list, dtype=float)
    mag = np.asarray(mag_list, dtype=float)
    phase = np.asarray(phase_list, dtype=float)

    # 周波数昇順にそろえる
    order = np.argsort(f)
    f, mag, phase = f[order], mag[order], phase[order]

    # 同一周波数がある場合は最初の値だけ残す
    unique_f, idx = np.unique(f, return_index=True)
    f, mag, phase = unique_f, mag[idx], phase[idx]

    return CsvData(path, f, mag, phase, header_row, freq_col, mag_col, phase_col)


# -----------------------------------------------------------------------------
# 補正・フィッティング
# -----------------------------------------------------------------------------
def fit_equivalent_model(
    freq_hz: np.ndarray,
    z_target: np.ndarray,
    rs_mode: str = RS_MODE_DEFAULT,
    rs_min_ohm: float = RS_MIN_OHM_DEFAULT,
    rs_fixed_ohm: float = RS_FIXED_OHM_DEFAULT,
) -> FitResult:
    """
    等価回路をフィットする。

    Rs制約:
        rs_mode="free"
            従来どおり Rs も自由にフィットする。
        rs_mode="lower_bound"
            Rsをフィットするが、Rs >= rs_min_ohm とする。
            全周波数Zshort補正で Rs が 10 uΩ などに張り付く場合の推奨設定。
        rs_mode="fixed"
            Rs = rs_fixed_ohm に固定し、Lcm/Cp/Rp/R2/L2/C2だけをフィットする。
            DMMや4端子測定のDC抵抗を優先したい場合に使う。
    """
    if least_squares is None:
        raise RuntimeError(
            "scipy が見つかりません。コマンドプロンプトで `pip install scipy` を実行してください。"
        ) from _SCIPY_IMPORT_ERROR

    f = np.asarray(freq_hz, dtype=float)
    zt = np.asarray(z_target, dtype=complex)
    mag = np.abs(zt)
    phase = np.rad2deg(np.angle(zt))
    w = 2.0 * np.pi * f

    # 初期値推定
    low_mask = (f <= min(1.0e4, f.max() / 100.0)) & (np.imag(zt) > 0)
    if np.sum(low_mask) < 5:
        low_mask = (f <= np.percentile(f, 20)) & (np.imag(zt) > 0)
    if np.sum(low_mask) < 3:
        low_mask = np.arange(len(f)) < max(3, len(f) // 10)

    Rs0 = float(np.nanmedian(np.maximum(np.real(zt[low_mask]), 1e-4)))
    L0 = float(np.nanmedian(np.imag(zt[low_mask]) / w[low_mask]))
    if not np.isfinite(Rs0) or Rs0 <= 0:
        Rs0 = 0.2
    if not np.isfinite(L0) or L0 <= 0:
        L0 = 1e-3
    rs_initial_ohm = float(Rs0)

    # 主自己共振ピーク
    peak_mask = f > max(f.min() * 5.0, 1.0e3)
    if not np.any(peak_mask):
        peak_mask = np.ones_like(f, dtype=bool)
    idx_peak_rel = int(np.argmax(mag[peak_mask]))
    idx_peak = np.where(peak_mask)[0][idx_peak_rel]
    f_peak = float(f[idx_peak])
    peak_mag = float(mag[idx_peak])
    Cp0 = 1.0 / ((2.0 * np.pi * f_peak) ** 2 * L0)
    Rp0 = max(peak_mag, 100.0)

    # 10〜25MHz付近の谷。範囲外なら高周波側1/3から推定。
    hf_mask = (f >= 1.0e7) & (f <= 2.5e7)
    if np.sum(hf_mask) < 5:
        hf_mask = f >= np.percentile(f, 70)
    idx_min_rel = int(np.argmin(mag[hf_mask]))
    idx_min = np.where(hf_mask)[0][idx_min_rel]
    f_hf = float(f[idx_min])
    min_mag = float(mag[idx_min])

    R20 = max(min_mag, 0.1)
    L20 = 2.0e-6
    C20 = 1.0 / ((2.0 * np.pi * f_hf) ** 2 * L20)

    # パラメータ順: Rs, Lcm, Cp, Rp, R2, L2, C2
    p0 = np.array([Rs0, L0, Cp0, Rp0, R20, L20, C20], dtype=float)
    lb = np.array([1e-5, 1e-7, 1e-14, 1e1, 1e-3, 1e-9, 1e-14], dtype=float)
    ub = np.array([1e2, 1e0, 1e-8, 1e8, 1e6, 1e-2, 1e-6], dtype=float)

    rs_mode_effective = (rs_mode or "free").strip().lower()
    if rs_mode_effective not in ("free", "lower_bound", "fixed"):
        rs_mode_effective = "free"

    try:
        rs_min_ohm = float(rs_min_ohm)
    except Exception:
        rs_min_ohm = RS_MIN_OHM_DEFAULT
    try:
        rs_fixed_ohm = float(rs_fixed_ohm)
    except Exception:
        rs_fixed_ohm = RS_FIXED_OHM_DEFAULT

    if not np.isfinite(rs_min_ohm) or rs_min_ohm <= 0:
        rs_min_ohm = RS_MIN_OHM_DEFAULT
    if not np.isfinite(rs_fixed_ohm) or rs_fixed_ohm <= 0:
        rs_fixed_ohm = RS_FIXED_OHM_DEFAULT

    active_mask = np.ones_like(p0, dtype=bool)
    rs_note = "Rs自由フィット。従来互換。"

    if rs_mode_effective == "lower_bound":
        lb[0] = max(lb[0], rs_min_ohm)
        p0[0] = max(p0[0], lb[0] * 1.01)
        rs_note = f"Rs下限制約: Rs >= {lb[0]:.6g} Ω"
    elif rs_mode_effective == "fixed":
        # 固定値が既定範囲外でも使えるよう、Rsだけ範囲を広げる。
        lb[0] = min(lb[0], rs_fixed_ohm)
        ub[0] = max(ub[0], rs_fixed_ohm)
        p0[0] = rs_fixed_ohm
        active_mask[0] = False
        rs_note = f"Rs固定: Rs = {rs_fixed_ohm:.6g} Ω"

    p0 = np.clip(p0, lb * 1.01, ub / 1.01)
    if rs_mode_effective == "fixed":
        p0[0] = rs_fixed_ohm

    # 重み: 主共振と高周波谷付近を少し重視
    logf = np.log10(f)
    weights = np.ones_like(f)
    weights += 1.5 * np.exp(-0.5 * ((logf - np.log10(f_peak)) / 0.12) ** 2)
    weights += 1.0 * np.exp(-0.5 * ((logf - np.log10(f_hf)) / 0.10) ** 2)

    def _dict_from_vals(vals: np.ndarray) -> dict:
        return {
            "Rs": vals[0],
            "Lcm": vals[1],
            "Cp": vals[2],
            "Rp": vals[3],
            "R2": vals[4],
            "L2": vals[5],
            "C2": vals[6],
        }

    def unpack(logp_active: np.ndarray) -> dict:
        vals = p0.copy()
        vals[active_mask] = np.exp(logp_active)
        if rs_mode_effective == "fixed":
            vals[0] = rs_fixed_ohm
        return _dict_from_vals(vals)

    target_mag_db = 20.0 * np.log10(np.maximum(np.abs(zt), 1e-30))
    target_phase = np.rad2deg(np.angle(zt))

    def residual(logp_active: np.ndarray) -> np.ndarray:
        p = unpack(logp_active)
        zm = model_impedance(f, p)
        model_mag_db = 20.0 * np.log10(np.maximum(np.abs(zm), 1e-30))
        model_phase = np.rad2deg(np.angle(zm))
        r_mag = (model_mag_db - target_mag_db) * weights
        r_phase = wrap_deg(model_phase - target_phase) / 3.0 * weights
        r = np.r_[r_mag, r_phase]
        if not np.all(np.isfinite(r)):
            return np.full_like(r, 1e30, dtype=float)
        return r

    p0_active = p0[active_mask]
    lb_active = lb[active_mask]
    ub_active = ub[active_mask]

    result = least_squares(
        residual,
        np.log(p0_active),
        bounds=(np.log(lb_active), np.log(ub_active)),
        loss="soft_l1",
        f_scale=1.0,
        max_nfev=5000,
    )
    p = unpack(result.x)
    zm = model_impedance(f, p)

    mag_err = 20.0 * np.log10(np.maximum(np.abs(zm), 1e-30) / np.maximum(np.abs(zt), 1e-30))
    phase_err = wrap_deg(np.rad2deg(np.angle(zm)) - np.rad2deg(np.angle(zt)))
    rms_mag = float(np.sqrt(np.mean(mag_err**2)))
    rms_phase = float(np.sqrt(np.mean(phase_err**2)))

    return FitResult(
        params=p,
        rms_mag_error_db=rms_mag,
        rms_phase_error_deg=rms_phase,
        f_main_peak_hz=f_peak,
        f_hf_min_hz=f_hf,
        rs_mode=rs_mode_effective,
        rs_min_ohm=rs_min_ohm,
        rs_fixed_ohm=rs_fixed_ohm,
        rs_initial_ohm=rs_initial_ohm,
        rs_lower_bound_ohm=float(lb[0]),
        rs_note=rs_note,
    )

def _estimate_lowfreq_short_rl(freq_hz: np.ndarray, zshort: np.ndarray) -> tuple[float, float, str]:
    """
    Zshort実測値から、低周波側の一定R/Lを推定する。
    目的は、20MHz付近の分布的なZshortをそのまま引かず、
    低周波で比較的素直な直列残留R/Lだけを補正に使うこと。
    """
    f = np.asarray(freq_hz, dtype=float)
    z = np.asarray(zshort, dtype=complex)

    # まず 100 kHz〜1 MHz を優先。点数不足なら最低側20%を使う。
    mask = (f >= 1.0e5) & (f <= 1.0e6) & np.isfinite(np.real(z)) & np.isfinite(np.imag(z))
    if np.sum(mask) < 5:
        limit = np.percentile(f, 20)
        mask = (f <= limit) & np.isfinite(np.real(z)) & np.isfinite(np.imag(z))

    if np.sum(mask) < 3:
        # 最後の保険
        mask = np.isfinite(np.real(z)) & np.isfinite(np.imag(z))

    rr = np.real(z[mask])
    ll = np.imag(z[mask]) / (2.0 * np.pi * f[mask])
    rr = rr[np.isfinite(rr)]
    ll = ll[np.isfinite(ll)]

    r_est = float(np.nanmedian(rr)) if rr.size else 0.0
    l_est = float(np.nanmedian(ll)) if ll.size else 0.0

    # 過補正防止のため、明らかな負値は0に丸める
    r_est = max(r_est, 0.0)
    l_est = max(l_est, 0.0)

    if np.any(mask):
        note = f"{float(np.nanmin(f[mask])):.6g}〜{float(np.nanmax(f[mask])):.6g} Hz の中央値"
    else:
        note = "推定不可のため0"

    return r_est, l_est, note


def analyze(
    meas_csv: str | Path,
    open_csv: str | Path,
    td_ns: float,
    rref_ohm: float = RREF_OHM_DEFAULT,
    short_csv: str | Path | None = None,
    short_mode: str = "open_only",
    rs_mode: str = RS_MODE_DEFAULT,
    rs_min_ohm: float = RS_MIN_OHM_DEFAULT,
    rs_fixed_ohm: float = RS_FIXED_OHM_DEFAULT,
) -> dict:
    meas = load_ad2_csv(meas_csv)
    opn = load_ad2_csv(open_csv)
    sht = load_ad2_csv(short_csv) if short_csv else None

    f = meas.frequency_hz
    Hmeas = complex_from_mag_phase(meas.mag_db, meas.phase_deg)
    Hopen_raw = complex_from_mag_phase(opn.mag_db, opn.phase_deg)
    Hopen = interp_complex_logf(opn.frequency_hz, Hopen_raw, f)

    if sht is not None:
        Hshort_raw = complex_from_mag_phase(sht.mag_db, sht.phase_deg)
        Hshort = interp_complex_logf(sht.frequency_hz, Hshort_raw, f)
    else:
        Hshort = np.zeros_like(Hmeas, dtype=complex)

    Zraw = h_to_impedance(Hmeas, rref_ohm)
    Zopen = h_to_impedance(Hopen, rref_ohm)
    Zshort_meas = h_to_impedance(Hshort, rref_ohm) if sht is not None else np.zeros_like(Zraw)

    rshort_est = 0.0
    lshort_est = 0.0
    short_est_note = "Zshort未選択"
    short_mode = (short_mode or "open_only").strip()

    # Zshortが選択されている場合は、補正モードに関係なく低周波一定R/Lを推定する。
    # v11ではlowfreq_rlモード時だけ推定していたため、fullモードや診断モードでは
    # 概要シートの推定Rshort/Lshortが0表示になっていた。
    if sht is not None:
        rshort_est, lshort_est, short_est_note = _estimate_lowfreq_short_rl(f, Zshort_meas)

    if sht is None:
        Zshort_eff = np.zeros_like(Zraw)
        correction_mode = "Open補正のみ"
        short_mode_effective = "open_only"
    elif short_mode == "lowfreq_rl":
        Zshort_eff = rshort_est + 1j * 2.0 * np.pi * f * lshort_est
        correction_mode = "Open+低周波一定R/L補正"
        short_mode_effective = "lowfreq_rl"
    elif short_mode == "full":
        # 参考モード。20MHz付近では過補正になることがあるため、デフォルトでは使わない。
        # ただし、7cm接続のようにZshortが十分小さい場合は比較用として有用。
        Zshort_eff = Zshort_meas
        correction_mode = "Open+全周波数Zshort補正（参考・非推奨）"
        short_mode_effective = "full"
    else:
        # Zshortを読み込んでも補正には使わない。診断値だけExcelへ出す。
        Zshort_eff = np.zeros_like(Zraw)
        correction_mode = "Open補正のみ（Zshortは診断用）"
        short_mode_effective = "open_only"

    # Open/Short補正
    #
    # Zshort_effは、選択モードに応じて
    #   0
    #   低周波一定R/L
    #   全周波数Zshort実測値
    # のいずれか。
    #
    # 補正式:
    #   Ydut = 1/(Zmeas - Zshort_eff) - 1/(Zopen - Zshort_eff)
    #
    # Zshort_eff=0なら従来のOpen補正と同じ。
    eps = 1e-30
    Zmeas_for_y = Zraw - Zshort_eff
    Zopen_for_y = Zopen - Zshort_eff

    Ymeas = 1.0 / (Zmeas_for_y + eps)
    Yopen = 1.0 / (Zopen_for_y + eps)
    Ydut = Ymeas - Yopen
    Zcorr = 1.0 / (Ydut + eps)

    td_s = td_ns * 1e-9
    tau_deg = 360.0 * f * td_s
    zcorr_mag = np.abs(Zcorr)
    zcorr_phase_deg = np.rad2deg(np.angle(Zcorr))
    zcorr_phase_tau_deg = wrap_deg(zcorr_phase_deg + tau_deg)

    # フィット対象は、補正後測定値にτ補正を加えたDUT端子面の複素インピーダンス。
    # モデル自体にはτ補正を加えない。
    Ztarget = zcorr_mag * np.exp(1j * np.deg2rad(zcorr_phase_tau_deg))
    fit = fit_equivalent_model(f, Ztarget, rs_mode=rs_mode, rs_min_ohm=rs_min_ohm, rs_fixed_ohm=rs_fixed_ohm)
    model_parts = decompose_model(f, fit.params)
    Zmodel = model_parts["Ztotal"]

    copen_pF = np.imag(Yopen) / (2.0 * np.pi * f) * 1e12
    ls_from_x_mH = np.imag(Zcorr) / (2.0 * np.pi * f) * 1e3
    lshort_meas_nH = np.imag(Zshort_meas) / (2.0 * np.pi * f) * 1e9
    lshort_eff_nH = np.imag(Zshort_eff) / (2.0 * np.pi * f) * 1e9
    phase_err = wrap_deg(np.rad2deg(np.angle(Zmodel)) - zcorr_phase_tau_deg)

    return {
        "meas": meas,
        "open": opn,
        "short": sht,
        "short_mode": short_mode_effective,
        "correction_mode": correction_mode,
        "short_est_note": short_est_note,
        "Rshort_est_ohm": rshort_est,
        "Lshort_est_H": lshort_est,
        "f": f,
        "Hmeas": Hmeas,
        "Hopen": Hopen,
        "Hshort": Hshort,
        "Zraw": Zraw,
        "Zopen": Zopen,
        "Zshort_meas": Zshort_meas,
        "Zshort_effective": Zshort_eff,
        # 互換用: 以降の処理で「実際に補正に使ったZshort」として扱う
        "Zshort": Zshort_eff,
        "Zmeas_for_y": Zmeas_for_y,
        "Zopen_for_y": Zopen_for_y,
        "Ymeas": Ymeas,
        "Yopen": Yopen,
        "Ydut": Ydut,
        "Zcorr": Zcorr,
        "Zcorr_phase_tau_deg": zcorr_phase_tau_deg,
        "tau_deg": tau_deg,
        "Copen_pF": copen_pF,
        "Lshort_meas_nH": lshort_meas_nH,
        "Lshort_eff_nH": lshort_eff_nH,
        "Ls_from_X_mH": ls_from_x_mH,
        "fit": fit,
        "model_parts": model_parts,
        "Zmodel": Zmodel,
        "phase_error": phase_err,
        "td_ns": td_ns,
        "rref_ohm": rref_ohm,
        "rs_mode": fit.rs_mode,
        "rs_min_ohm": fit.rs_min_ohm,
        "rs_fixed_ohm": fit.rs_fixed_ohm,
        "rs_note": fit.rs_note,
    }


# -----------------------------------------------------------------------------
# Excel出力
# -----------------------------------------------------------------------------
def _style_header(ws, row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(bottom=Side(style="thin", color="808080"))
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")


def _autofit_rough(ws, max_width: int = 32) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = 10
        for cell in col_cells[:200]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _add_scatter_chart(
    ws_chart,
    ws_data,
    title: str,
    description: str,
    x_col: int,
    y_cols: List[Tuple[int, str]],
    min_row: int,
    max_row: int,
    y_title: str,
    anchor: str = "A4",
    log_y: bool = False,
    x_min: float | None = None,
    x_max: float | None = None,
) -> None:
    ws_chart["A1"] = title
    ws_chart["A1"].font = Font(size=14, bold=True)
    ws_chart["A2"] = description
    ws_chart["A2"].alignment = Alignment(wrap_text=True)
    ws_chart.column_dimensions["A"].width = 110

    chart = ScatterChart()
    chart.title = title
    chart.style = 13
    chart.height = 14
    chart.width = 26
    chart.scatterStyle = "lineMarker"
    chart.x_axis.title = "Frequency [Hz]"
    chart.y_axis.title = y_title
    chart.x_axis.scaling.logBase = 10
    if x_min is not None:
        chart.x_axis.scaling.min = x_min
    if x_max is not None:
        chart.x_axis.scaling.max = x_max
    if log_y:
        chart.y_axis.scaling.logBase = 10

    xvalues = Reference(ws_data, min_col=x_col, min_row=min_row, max_row=max_row)
    for y_col, label in y_cols:
        yvalues = Reference(ws_data, min_col=y_col, min_row=min_row, max_row=max_row)
        series = Series(yvalues, xvalues, title=label)
        series.marker.symbol = "circle"
        series.marker.size = 3
        chart.series.append(series)
    chart.legend.position = "b"
    ws_chart.add_chart(chart, anchor)



def _sig_text(value: float, digits: int = 3) -> str:
    """軸ラベル用の有効数字文字列。"""
    if not np.isfinite(value):
        return ""
    if abs(value) >= 100:
        s = f"{value:.0f}"
    elif abs(value) >= 10:
        s = f"{value:.1f}"
    elif abs(value) >= 1:
        s = f"{value:.2f}"
    else:
        s = f"{value:.3g}"
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s


def _format_hz(x: float, _pos=None) -> str:
    if not np.isfinite(x) or x <= 0:
        return ""
    if x >= 1e6:
        return f"{_sig_text(x / 1e6)} MHz"
    if x >= 1e3:
        return f"{_sig_text(x / 1e3)} kHz"
    return f"{_sig_text(x)} Hz"


def _format_ohm(x: float, _pos=None) -> str:
    if not np.isfinite(x) or x <= 0:
        return ""
    if x >= 1e6:
        return f"{_sig_text(x / 1e6)} MΩ"
    if x >= 1e3:
        return f"{_sig_text(x / 1e3)} kΩ"
    return f"{_sig_text(x)} Ω"


def _format_deg(x: float, _pos=None) -> str:
    if not np.isfinite(x):
        return ""
    return f"{_sig_text(x)}°"


def _format_pf(x: float, _pos=None) -> str:
    if not np.isfinite(x):
        return ""
    return f"{_sig_text(x)} pF"




def _log_ticks(vmin: float, vmax: float) -> List[float]:
    """ログ軸に出す主要目盛。広い範囲はdecade、拡大図は1-2-5を表示。"""
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmin <= 0 or vmax <= vmin:
        return []
    span_dec = math.log10(vmax / vmin)
    mults = [1.0] if span_dec > 3.0 else [1.0, 2.0, 5.0]
    e0 = math.floor(math.log10(vmin)) - 1
    e1 = math.ceil(math.log10(vmax)) + 1
    ticks = []
    for e in range(e0, e1 + 1):
        base = 10.0 ** e
        for m in mults:
            t = m * base
            if vmin <= t <= vmax:
                ticks.append(t)
    return sorted(set(ticks))

def _prepare_plot_font() -> None:
    """Windowsで日本語が出やすいフォント候補を設定。"""
    if plt is None:
        return
    candidates = ["Noto Sans CJK JP", "Meiryo", "Yu Gothic", "MS Gothic", "DejaVu Sans"]
    try:
        import matplotlib.font_manager as fm
        installed = {f.name for f in fm.fontManager.ttflist}
        family = next((name for name in candidates if name in installed), "DejaVu Sans")
    except Exception:
        family = candidates
    plt.rcParams["font.family"] = family
    plt.rcParams["axes.unicode_minus"] = False


def _add_plot_image_sheet(
    ws_chart,
    title: str,
    description: str,
    png_path: Path,
    freq_hz: np.ndarray,
    series_list: List[dict],
    y_title: str,
    y_formatter,
    log_y: bool = False,
    x_min: float | None = None,
    x_max: float | None = None,
    y_min: float | None = None,
    y_max: float | None = None,
) -> None:
    """7/1版に近い見やすいMatplotlib画像を作り、Excelシートへ貼り付ける。"""
    if plt is None:
        raise RuntimeError(
            "matplotlib が見つかりません。コマンドプロンプトで `pip install matplotlib pillow` を実行してください。"
        ) from _MATPLOTLIB_IMPORT_ERROR

    _prepare_plot_font()

    f = np.asarray(freq_hz, dtype=float)
    xmin = float(x_min) if x_min is not None else float(np.nanmin(f[f > 0]))
    xmax = float(x_max) if x_max is not None else float(np.nanmax(f))
    xmin = max(xmin, float(np.nanmin(f[f > 0])))
    xmax = min(xmax, float(np.nanmax(f)))

    fig, ax = plt.subplots(figsize=(12.8, 6.8))
    all_y_for_ticks = []
    for idx, item in enumerate(series_list):
        y = np.asarray(item["y"], dtype=float)
        mask = np.isfinite(f) & np.isfinite(y) & (f > 0) & (f >= xmin) & (f <= xmax)
        if log_y:
            mask &= y > 0
        if not np.any(mask):
            continue
        style = item.get("style", "line")
        label = item.get("label", f"Series {idx+1}")
        all_y_for_ticks.append(y[mask])
        if style == "meas":
            ax.plot(
                f[mask], y[mask],
                linestyle="None", marker="o", markersize=3.0,
                markeredgewidth=0.0, alpha=0.75, label=label,
            )
        elif style == "open":
            ax.plot(f[mask], y[mask], linewidth=1.8, alpha=0.9, label=label)
        else:
            ax.plot(f[mask], y[mask], linewidth=2.0, label=label)

    ax.set_xscale("log")
    if log_y:
        ax.set_yscale("log")
    ax.set_xlim(xmin, xmax)
    x_ticks = _log_ticks(xmin, xmax)
    if x_ticks:
        ax.set_xticks(x_ticks)
    if y_min is not None or y_max is not None:
        ax.set_ylim(y_min, y_max)
    if log_y and all_y_for_ticks:
        yy = np.concatenate(all_y_for_ticks)
        yy = yy[np.isfinite(yy) & (yy > 0)]
        if yy.size:
            ymin_tick = y_min if y_min is not None else float(np.nanmin(yy))
            ymax_tick = y_max if y_max is not None else float(np.nanmax(yy))
            y_ticks = _log_ticks(ymin_tick, ymax_tick)
            if y_ticks:
                ax.set_yticks(y_ticks)

    ax.set_title(title, fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("Frequency [Hz]", fontsize=12)
    ax.set_ylabel(y_title, fontsize=12)
    ax.xaxis.set_major_formatter(FuncFormatter(_format_hz))
    ax.xaxis.set_minor_formatter(NullFormatter())
    ax.yaxis.set_major_formatter(FuncFormatter(y_formatter))
    if log_y:
        ax.yaxis.set_minor_formatter(NullFormatter())
    ax.tick_params(axis="both", which="major", labelsize=10)
    ax.grid(True, which="major", linewidth=0.8, alpha=0.35)
    ax.grid(True, which="minor", linewidth=0.5, alpha=0.18)
    ax.legend(loc="best", fontsize=10, framealpha=0.9)
    fig.tight_layout()
    fig.savefig(png_path, dpi=160, bbox_inches="tight")
    plt.close(fig)

    ws_chart.sheet_view.showGridLines = False
    ws_chart["A1"] = title
    ws_chart["A1"].font = Font(size=14, bold=True)
    ws_chart["A2"] = description
    ws_chart["A2"].alignment = Alignment(wrap_text=True)
    ws_chart["A3"] = f"横軸: Frequency [Hz] / 縦軸: {y_title}"
    ws_chart["A3"].font = Font(bold=True, color="404040")
    ws_chart.column_dimensions["A"].width = 120
    ws_chart.row_dimensions[1].height = 22
    ws_chart.row_dimensions[2].height = 36
    ws_chart.row_dimensions[3].height = 20

    img = XLImage(str(png_path))
    img.width = 1180
    img.height = 640
    ws_chart.add_image(img, "A5")


def _draw_resistor(ax, x1, x2, y, zigzags=6, amp=0.12, lw=2.4, color='black'):
    xs = [x1]
    ys = [y]
    length = x2 - x1
    step = length / (zigzags * 2)
    x = x1
    sign = 1
    for _ in range(zigzags * 2 - 1):
        x += step
        xs.append(x)
        ys.append(y + amp * sign)
        sign *= -1
    xs.append(x2)
    ys.append(y)
    ax.plot(xs, ys, color=color, lw=lw)


def _draw_inductor(ax, x1, x2, y, loops=4, radius=0.12, lw=2.4, color='black'):
    import numpy as _np
    ax.plot([x1, x1 + radius], [y, y], color=color, lw=lw)
    span = x2 - x1 - 2 * radius
    pitch = span / loops
    start = x1 + radius
    for i in range(loops):
        t = _np.linspace(0, _np.pi, 60)
        xc = start + i * pitch + pitch / 2
        r = pitch / 2
        xx = xc - r * _np.cos(t)
        yy = y + radius * _np.sin(t)
        ax.plot(xx, yy, color=color, lw=lw)
    ax.plot([x2 - radius, x2], [y, y], color=color, lw=lw)


def _draw_capacitor_h(ax, x1, x2, y, gap=0.10, plate=0.38, lw=2.4, color='black'):
    xm = (x1 + x2) / 2
    ax.plot([x1, xm - gap], [y, y], color=color, lw=lw)
    ax.plot([xm - gap, xm - gap], [y - plate/2, y + plate/2], color=color, lw=lw)
    ax.plot([xm + gap, xm + gap], [y - plate/2, y + plate/2], color=color, lw=lw)
    ax.plot([xm + gap, x2], [y, y], color=color, lw=lw)


def _draw_resistor_v(ax, x, y1, y2, zigzags=6, amp=0.16, lw=2.4, color='black'):
    ys = [y1]
    xs = [x]
    length = y2 - y1
    step = length / (zigzags * 2)
    y = y1
    sign = 1
    for _ in range(zigzags * 2 - 1):
        y += step
        ys.append(y)
        xs.append(x + amp * sign)
        sign *= -1
    ys.append(y2)
    xs.append(x)
    ax.plot(xs, ys, color=color, lw=lw)


def _draw_inductor_v(ax, x, y1, y2, loops=4, radius=0.16, lw=2.4, color='black'):
    import numpy as _np
    ax.plot([x, x], [y1, y1 + radius], color=color, lw=lw)
    span = y2 - y1 - 2 * radius
    pitch = span / loops
    start = y1 + radius
    for i in range(loops):
        t = _np.linspace(0, _np.pi, 60)
        yc = start + i * pitch + pitch / 2
        r = pitch / 2
        yy = yc - r * _np.cos(t)
        xx = x + radius * _np.sin(t)
        ax.plot(xx, yy, color=color, lw=lw)
    ax.plot([x, x], [y2 - radius, y2], color=color, lw=lw)


def _draw_capacitor_v(ax, x, y1, y2, gap=0.12, plate=0.56, lw=2.4, color='black'):
    ym = (y1 + y2) / 2
    ax.plot([x, x], [y1, ym - gap], color=color, lw=lw)
    ax.plot([x - plate/2, x + plate/2], [ym - gap, ym - gap], color=color, lw=lw)
    ax.plot([x - plate/2, x + plate/2], [ym + gap, ym + gap], color=color, lw=lw)
    ax.plot([x, x], [ym + gap, y2], color=color, lw=lw)


def _draw_label_right(ax, x, y, name, value_text, color='black', name_size=15, value_size=13, dx=0.55):
    ax.text(x + dx, y + 0.16, name, ha='left', va='bottom', fontsize=name_size, fontweight='bold', color=color)
    ax.text(x + dx, y - 0.16, value_text, ha='left', va='top', fontsize=value_size, color=color)


def _draw_equivalent_circuit_image(p: dict, png_path: Path) -> None:
    """等価回路の説明用画像を作成する。図記号を90度回転し、P/Nの配線が重ならないレイアウトにする。"""
    if plt is None:
        raise RuntimeError(
            "matplotlib が見つかりません。コマンドプロンプトで `pip install matplotlib pillow` を実行してください。"
        ) from _MATPLOTLIB_IMPORT_ERROR

    from matplotlib.patches import FancyBboxPatch

    _prepare_plot_font()

    def fmt_eng(v: float, unit: str) -> str:
        av = abs(v)
        table = [
            (1e9, 'G'), (1e6, 'M'), (1e3, 'k'),
            (1, ''), (1e-3, 'm'), (1e-6, 'u'), (1e-9, 'n'), (1e-12, 'p')
        ]
        for scale, prefix in table:
            if av >= scale * 0.999:
                return f'{v/scale:.4g} {prefix}{unit}'
        return f'{v:.4g} {unit}'

    col_low = '#303030'
    col_main = '#2F75B5'
    col_hf = '#E67E22'
    fill_low = '#F3F3F3'
    fill_main = '#EAF3FF'
    fill_hf = '#FFF0E2'

    fig, ax = plt.subplots(figsize=(15.8, 8.6))
    ax.set_xlim(0, 17)
    ax.set_ylim(0, 12.0)
    ax.axis('off')

    top_y = 7.45
    bot_y = 1.65
    left_x = 0.9
    right_x = 16.1
    branch_x = [2.6, 6.8, 10.2, 13.9]

    ax.text(0.35, 11.25, '等価回路モデル（4並列枝）', fontsize=21, fontweight='bold')
    ax.text(0.35, 10.58, 'Z = 1 / ( 1/(Rs+sLcm) + 1/Rp + sCp + 1/(R2+sL2+1/(sC2)) )', fontsize=13)
    ax.text(0.35, 10.02, '青 = 主自己共振枝（Rp, Cp） / 橙 = 20 MHz枝（R2, L2, C2）', fontsize=12)

    boxes = [
        (1.35, 0.9, 2.5, 7.25, fill_low, col_low, '低周波枝'),
        (5.45, 0.9, 5.9, 7.25, fill_main, col_main, '主自己共振枝'),
        (12.45, 0.9, 2.9, 7.25, fill_hf, col_hf, '20 MHz枝'),
    ]
    for x, y, w, h, fc, ec, title in boxes:
        patch = FancyBboxPatch((x, y), w, h, boxstyle='round,pad=0.18,rounding_size=0.18',
                               linewidth=2.0, edgecolor=ec, facecolor=fc, zorder=0)
        ax.add_patch(patch)
        ax.text(x + w/2, y + h - 0.36, title, ha='center', va='center', fontsize=14, fontweight='bold', color=ec)

    ax.plot([left_x, right_x], [top_y, top_y], color='black', lw=2.8)
    ax.plot([left_x, right_x], [bot_y, bot_y], color='black', lw=2.8)
    ax.plot(left_x, top_y, 'ko', ms=7)
    ax.plot(left_x, bot_y, 'ko', ms=7)
    ax.plot(right_x, top_y, 'ko', ms=7)
    ax.plot(right_x, bot_y, 'ko', ms=7)
    ax.text(left_x - 0.16, top_y + 0.45, 'P', fontsize=14, fontweight='bold')
    ax.text(left_x - 0.16, bot_y - 0.78, 'N', fontsize=14, fontweight='bold')
    ax.text(right_x + 0.08, top_y + 0.45, 'P', fontsize=14, fontweight='bold')
    ax.text(right_x + 0.08, bot_y - 0.78, 'N', fontsize=14, fontweight='bold')

    x = branch_x[0]
    ax.plot([x, x], [top_y, 6.95], color=col_low, lw=2.4)
    _draw_resistor_v(ax, x, 5.75, 6.95, lw=2.6, color=col_low)
    ax.plot([x, x], [4.95, 5.75], color=col_low, lw=2.4)
    _draw_inductor_v(ax, x, 3.45, 4.95, lw=2.6, color=col_low)
    ax.plot([x, x], [bot_y, 3.45], color=col_low, lw=2.4)
    _draw_label_right(ax, x, 6.35, 'Rs', fmt_eng(p['Rs'], 'Ω'), color=col_low, dx=0.62)
    _draw_label_right(ax, x, 4.2, 'Lcm', fmt_eng(p['Lcm'], 'H'), color=col_low, dx=0.62)

    x = branch_x[1]
    ax.plot([x, x], [top_y, 6.95], color=col_main, lw=2.5)
    _draw_resistor_v(ax, x, 4.15, 6.95, lw=2.8, color=col_main)
    ax.plot([x, x], [bot_y, 4.15], color=col_main, lw=2.5)
    _draw_label_right(ax, x, 5.55, 'Rp', fmt_eng(p['Rp'], 'Ω'), color=col_main, dx=0.68)

    x = branch_x[2]
    ax.plot([x, x], [top_y, 6.05], color=col_main, lw=2.5)
    _draw_capacitor_v(ax, x, 3.95, 6.05, gap=0.16, plate=0.72, lw=2.8, color=col_main)
    ax.plot([x, x], [bot_y, 3.95], color=col_main, lw=2.5)
    _draw_label_right(ax, x, 5.0, 'Cp', fmt_eng(p['Cp'], 'F'), color=col_main, dx=0.68)

    x = branch_x[3]
    # 高さ方向を広げて、L2 と C2 が重ならないように間隔を確保
    ax.plot([x, x], [top_y, 7.25], color=col_hf, lw=2.5)
    _draw_resistor_v(ax, x, 6.05, 7.25, lw=2.8, color=col_hf)
    ax.plot([x, x], [5.75, 6.05], color=col_hf, lw=2.5)
    _draw_inductor_v(ax, x, 4.25, 5.75, lw=2.8, color=col_hf)
    ax.plot([x, x], [3.65, 4.25], color=col_hf, lw=2.5)
    _draw_capacitor_v(ax, x, 2.35, 3.65, gap=0.15, plate=0.72, lw=2.8, color=col_hf)
    ax.plot([x, x], [bot_y, 2.35], color=col_hf, lw=2.5)
    _draw_label_right(ax, x, 6.65, 'R2', fmt_eng(p['R2'], 'Ω'), color=col_hf, dx=0.68)
    _draw_label_right(ax, x, 5.0, 'L2', fmt_eng(p['L2'], 'H'), color=col_hf, dx=0.68)
    _draw_label_right(ax, x, 2.95, 'C2', fmt_eng(p['C2'], 'F'), color=col_hf, dx=0.68)

    ax.text(8.4, 0.68, '主自己共振の山は主に Lcm と Cp、損失は Rp で表現', ha='center', fontsize=12, color=col_main)
    ax.text(13.9, 0.22, '20 MHz付近の谷 / 高周波側共振を\n直列枝 R2-L2-C2 で表現', ha='center', fontsize=12, color=col_hf)

    fig.tight_layout()
    fig.savefig(png_path, dpi=180, bbox_inches='tight')
    plt.close(fig)

def export_excel(result: dict, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("概要")
    ws_data = wb.create_sheet("計算データ")
    ws_points = wb.create_sheet("主要点")
    ws_amp = wb.create_sheet("グラフ_振幅")
    ws_phase = wb.create_sheet("グラフ_位相_tau")
    ws_open = wb.create_sheet("グラフ_Open容量")
    ws_peak = wb.create_sheet("グラフ_主自己共振拡大")
    ws_hf = wb.create_sheet("グラフ_20MHz拡大")
    ws_param = wb.create_sheet("等価回路パラメータ")
    ws_spice = wb.create_sheet("LTspice_Subckt")

    f = result["f"]
    Hmeas = result["Hmeas"]
    Hopen = result["Hopen"]
    Zraw = result["Zraw"]
    Zopen = result["Zopen"]
    Ymeas = result["Ymeas"]
    Yopen = result["Yopen"]
    Ydut = result["Ydut"]
    Zcorr = result["Zcorr"]
    fit: FitResult = result["fit"]
    p = fit.params
    parts = result["model_parts"]
    Zmodel = result["Zmodel"]

    # 概要
    ws_summary["A1"] = "SC-05-30J Open/Short補正・τ補正・等価回路フィット 周波数特性"
    ws_summary["A1"].font = Font(size=14, bold=True)
    rows = [
        ("項目", "値", "単位", "備考"),
        ("測定CSV", str(result["meas"].path.name), "", "GUIで選択した測定データ"),
        ("Open CSV", str(result["open"].path.name), "", "GUIで選択したZopenデータ"),
        ("Short CSV", str(result["short"].path.name) if result.get("short") is not None else "未選択", "", "GUIで選択したZshortデータ。未選択時は従来のOpen補正のみ。"),
        ("補正モード", result.get("correction_mode", "Open補正のみ"), "", "デフォルトはOpen補正のみ。Zshortは診断用として読み込み可能。"),
        ("測定CSV列", f"freq={result["meas"].freq_col+1}, mag={result["meas"].mag_col+1}, phase={result["meas"].phase_col+1}", "列番号", "1始まり。Channel 2 Magnitude/Phaseになっていることを確認"),
        ("Open CSV列", f"freq={result["open"].freq_col+1}, mag={result["open"].mag_col+1}, phase={result["open"].phase_col+1}", "列番号", "1始まり。Channel 2 Magnitude/Phaseになっていることを確認"),
        ("Short CSV列", f"freq={result['short'].freq_col+1}, mag={result['short'].mag_col+1}, phase={result['short'].phase_col+1}" if result.get("short") is not None else "-", "列番号", "Zshort選択時のみ。Channel 2 Magnitude/Phaseになっていることを確認"),
        ("Rref", result["rref_ohm"], "Ω", "Z = Rref H/(1-H)"),
        ("td", result["td_ns"], "ns", "測定値だけに適用。モデル位相には適用しない。"),
        ("Rs制約モード", fit.rs_mode, "", fit.rs_note),
        ("Rs初期推定", fit.rs_initial_ohm, "Ω", "低周波側の実部から推定した初期値。制約前の参考値。"),
        ("Rs下限", fit.rs_lower_bound_ohm, "Ω", "lower_boundモードではこの値以上。freeモードでは従来下限。fixedモードでは参考。"),
        ("Rs固定値", fit.rs_fixed_ohm, "Ω", "fixedモード選択時に使用。"),
        ("Rsフィット結果", p["Rs"], "Ω", "LTspiceモデルのRs。"),
        ("Zshort補正の扱い", result.get("short_mode", "open_only"), "", "open_only=診断のみ / lowfreq_rl=低周波一定R/Lのみ補正 / full=全周波数差し引き参考"),
        ("推定Rshort", result.get("Rshort_est_ohm", 0.0), "Ω", result.get("short_est_note", "") + "。Zshort選択時は補正モードに関係なく表示。"),
        ("推定Lshort", result.get("Lshort_est_H", 0.0), "H", "低周波一定R/L補正モードで使用。fullモードでは参考値として表示。"),
        ("測定点数", len(f), "点", "周波数軸は測定CSVに合わせ、Zopenはlog f軸で補間"),
        ("主自己共振ピーク", fit.f_main_peak_hz, "Hz", "補正後+τ補正の|Z|ピーク付近"),
        ("高周波側谷", fit.f_hf_min_hz, "Hz", "10〜25MHz付近。範囲外なら高周波側から推定"),
        ("振幅RMS誤差", fit.rms_mag_error_db, "dB", "20log(|Zmodel|/|Zmeas|)"),
        ("位相RMS誤差", fit.rms_phase_error_deg, "deg", "測定値は補正後+τ補正、モデル位相はτ補正なし"),
    ]
    for r, row in enumerate(rows, 3):
        for c, val in enumerate(row, 1):
            ws_summary.cell(r, c, val)
    _style_header(ws_summary, 3, 1, 4)
    _autofit_rough(ws_summary, 45)

    # 計算データ
    headers = [
        "Frequency_Hz", "Meas_Ch2_Mag_dB", "Meas_Phase_deg", "Open_Ch2_Mag_dB_interp", "Open_Phase_deg_interp",
        "Short_Ch2_Mag_dB_interp", "Short_Phase_deg_interp",
        "Hmeas_Re", "Hmeas_Im", "Hopen_Re", "Hopen_Im", "Hshort_Re", "Hshort_Im",
        "Z_raw_R_Ohm", "Z_raw_X_Ohm", "Z_raw_mag_Ohm", "Z_raw_phase_deg",
        "Zopen_R_Ohm", "Zopen_X_Ohm",
        "Zshort_meas_R_Ohm", "Zshort_meas_X_Ohm", "Zshort_meas_mag_Ohm", "Lshort_meas_nH",
        "Zshort_used_R_Ohm", "Zshort_used_X_Ohm", "Zshort_used_mag_Ohm", "Lshort_used_nH",
        "Zmeas_minus_Zshort_used_R_Ohm", "Zmeas_minus_Zshort_used_X_Ohm",
        "Zopen_minus_Zshort_used_R_Ohm", "Zopen_minus_Zshort_used_X_Ohm",
        "Ymeas_Re_S", "Ymeas_Im_S", "Yopen_Re_S", "Yopen_Im_S", "Copen_pF",
        "Ydut_Re_S", "Ydut_Im_S", "Zcorr_R_Ohm", "Zcorr_X_Ohm", "Zcorr_mag_Ohm", "Zcorr_phase_deg",
        "Tau_correction_deg", "Zcorr_phase_tau_deg", "Ls_from_X_mH",
        "Model_omega_rad_s", "Model_Y1_Re_S", "Model_Y1_Im_S", "Model_Ycp_Im_S", "Model_Yrp_Re_S",
        "Model_X2_Ohm", "Model_Y2_Re_S", "Model_Y2_Im_S", "Model_Ytotal_Re_S", "Model_Ytotal_Im_S",
        "Zmodel_R_Ohm", "Zmodel_X_Ohm", "Zmodel_mag_Ohm", "Zmodel_phase_deg_no_tau", "Zmodel_mag_dBOhm",
        "Phase_error_vs_meas_tau_deg",
    ]
    ws_data.append(headers)
    _style_header(ws_data, 1, 1, len(headers))

    Hopen_mag = 20.0 * np.log10(np.maximum(np.abs(Hopen), 1e-30))
    Hopen_phase = np.rad2deg(np.angle(Hopen))
    Hshort = result.get("Hshort", np.zeros_like(Hmeas))
    Zshort_meas = result.get("Zshort_meas", np.zeros_like(Zraw))
    Zshort_eff = result.get("Zshort_effective", np.zeros_like(Zraw))
    Hshort_mag = 20.0 * np.log10(np.maximum(np.abs(Hshort), 1e-30))
    Hshort_phase = np.rad2deg(np.angle(Hshort))
    w = 2.0 * np.pi * f
    for i in range(len(f)):
        row = [
            f[i], result["meas"].mag_db[i], result["meas"].phase_deg[i], Hopen_mag[i], Hopen_phase[i],
            Hshort_mag[i], Hshort_phase[i],
            np.real(Hmeas[i]), np.imag(Hmeas[i]), np.real(Hopen[i]), np.imag(Hopen[i]), np.real(Hshort[i]), np.imag(Hshort[i]),
            np.real(Zraw[i]), np.imag(Zraw[i]), abs(Zraw[i]), np.rad2deg(np.angle(Zraw[i])),
            np.real(Zopen[i]), np.imag(Zopen[i]),
            np.real(Zshort_meas[i]), np.imag(Zshort_meas[i]), abs(Zshort_meas[i]), result["Lshort_meas_nH"][i],
            np.real(Zshort_eff[i]), np.imag(Zshort_eff[i]), abs(Zshort_eff[i]), result["Lshort_eff_nH"][i],
            np.real(result["Zmeas_for_y"][i]), np.imag(result["Zmeas_for_y"][i]),
            np.real(result["Zopen_for_y"][i]), np.imag(result["Zopen_for_y"][i]),
            np.real(Ymeas[i]), np.imag(Ymeas[i]), np.real(Yopen[i]), np.imag(Yopen[i]), result["Copen_pF"][i],
            np.real(Ydut[i]), np.imag(Ydut[i]), np.real(Zcorr[i]), np.imag(Zcorr[i]), abs(Zcorr[i]), np.rad2deg(np.angle(Zcorr[i])),
            result["tau_deg"][i], result["Zcorr_phase_tau_deg"][i], result["Ls_from_X_mH"][i],
            w[i], np.real(parts["Y1"][i]), np.imag(parts["Y1"][i]), np.imag(parts["Ycp"][i]), np.real(parts["Yrp"][i]),
            np.imag(parts["Z2"][i]), np.real(parts["Y2"][i]), np.imag(parts["Y2"][i]), np.real(parts["Ytotal"][i]), np.imag(parts["Ytotal"][i]),
            np.real(Zmodel[i]), np.imag(Zmodel[i]), abs(Zmodel[i]), np.rad2deg(np.angle(Zmodel[i])),
            20.0 * np.log10(max(abs(Zmodel[i]), 1e-30)), result["phase_error"][i],
        ]
        ws_data.append([float(x) if isinstance(x, (np.floating, np.integer)) else x for x in row])
    ws_data.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws_data.column_dimensions[get_column_letter(col)].width = 16
    for col in range(1, len(headers) + 1):
        for cell in ws_data[get_column_letter(col)][1:]:
            cell.number_format = "0.000E+00"

    # 主要点
    ws_points["A1"] = "主要周波数点比較"
    ws_points["A1"].font = Font(size=14, bold=True)
    point_headers = [
        "目標周波数[Hz]", "実周波数[Hz]", "|Z|補正前[Ω]", "|Z|補正後[Ω]",
        "位相補正前[deg]", "位相補正後[deg]", "τ補正後位相[deg]", "τ補正量[deg]",
        "|Zmodel|[Ω]", "モデル位相[deg]",
    ]
    ws_points.append([])
    ws_points.append(point_headers)
    _style_header(ws_points, 3, 1, len(point_headers))
    targets = [100, 1e3, 10e3, 100e3, fit.f_main_peak_hz, 1e6, 10e6, fit.f_hf_min_hz, 30e6]
    used = []
    for ft in targets:
        idx = int(np.argmin(np.abs(f - ft)))
        if idx in used:
            continue
        used.append(idx)
        ws_points.append([
            ft, f[idx], abs(Zraw[idx]), abs(Zcorr[idx]), np.rad2deg(np.angle(Zraw[idx])),
            np.rad2deg(np.angle(Zcorr[idx])), result["Zcorr_phase_tau_deg"][idx], result["tau_deg"][idx],
            abs(Zmodel[idx]), np.rad2deg(np.angle(Zmodel[idx])),
        ])
    _autofit_rough(ws_points, 24)

    # パラメータ
    ws_param["A1"] = "等価回路モデルパラメータ（今回測定値からフィット）"
    ws_param["A1"].font = Font(size=14, bold=True)
    param_rows = [
        ("素子/項目", "値", "単位", "説明"),
        ("Rs", p["Rs"], "Ω", "低周波側の直列抵抗。Rs制約: " + fit.rs_note),
        ("Rs_mode", fit.rs_mode, "", "free=従来 / lower_bound=下限制約 / fixed=固定"),
        ("Rs_initial", fit.rs_initial_ohm, "Ω", "制約前の初期推定値"),
        ("Lcm", p["Lcm"], "H", "低周波インダクタンス"),
        ("Cp", p["Cp"], "F", "主自己共振を表す並列容量"),
        ("Rp", p["Rp"], "Ω", "主自己共振ピークの損失抵抗"),
        ("R2", p["R2"], "Ω", "20MHz付近の谷を表す直列共振枝の抵抗"),
        ("L2", p["L2"], "H", "20MHz付近の直列共振枝インダクタンス"),
        ("C2", p["C2"], "F", "20MHz付近の直列共振枝容量"),
        ("f_main_peak_meas", fit.f_main_peak_hz, "Hz", "補正後+τ補正の振幅ピーク周波数"),
        ("f_HF_min_meas", fit.f_hf_min_hz, "Hz", "10〜25MHz範囲の最小|Z|周波数"),
        ("RMS_mag_error", fit.rms_mag_error_db, "dB", "全点の20log(|Zmodel|/|Zmeas|) RMS"),
        ("RMS_phase_error", fit.rms_phase_error_deg, "deg", "全点の位相差RMS。測定値はτ補正後、モデル位相はτ補正なし。"),
        ("", "", "", ""),
        ("回路形式", "Y = 1/(Rs + sLcm) + 1/Rp + sCp + 1/(R2 + sL2 + 1/(sC2))", "", "Z = 1/Y"),
        ("位相注記", "モデル位相にはτ補正を加えていません。位相グラフでは測定値の補正後+τ補正位相とモデルの素の位相を比較しています。", "", ""),
        ("Zshort注記", "Zshortを選択してもデフォルトでは補正に使わず診断値として出力します。低周波一定R/L補正時のみ推定Rshort/Lshortを使用し、fullモードは参考です。", "", ""),
    ]
    for r, row in enumerate(param_rows, 3):
        for c, val in enumerate(row, 1):
            ws_param.cell(r, c, val)
    _style_header(ws_param, 3, 1, 4)
    _autofit_rough(ws_param, 70)

    # 等価回路画像（直感的に分かるよう、パラメータ値入りの回路図を貼り付け）
    ws_param["A19"] = "画像化した等価回路"
    ws_param["A19"].font = Font(size=14, bold=True)
    ws_param["A19"].fill = PatternFill("solid", fgColor="EAF3FF")

    # LTspiceサブサーキット
    spice_lines = [
        "* SC-05-30J equivalent model generated by sc05_impedance_fit_gui.py",
        "* Model phase is not tau-corrected.",
        "* Topology matches: Y = 1/(Rs+sLcm) + 1/Rp + sCp + 1/(R2+sL2+1/(sC2))",
        ".subckt SC_05_30J_fit p n",
        f"Rs p a1 {p['Rs']:.10g}",
        f"Lcm a1 n {p['Lcm']:.10g}",
        f"Rp p n {p['Rp']:.10g}",
        f"Cp p n {p['Cp']:.10g}",
        f"R2 p b {p['R2']:.10g}",
        f"L2 b c {p['L2']:.10g}",
        f"C2 c n {p['C2']:.10g}",
        ".ends SC_05_30J_fit",
    ]
    for i, line in enumerate(spice_lines, 1):
        ws_spice.cell(i, 1, line)
    ws_spice.column_dimensions["A"].width = 90

    # グラフ（7/1版に近い画像貼り付け形式・5種）
    # Excelネイティブチャートではなく、Matplotlibで描画した画像を貼り付けることで、
    # 横軸/縦軸の数値単位表示を安定させる。
    plot_dir = Path(tempfile.mkdtemp(prefix="sc05_fit_plots_"))
    try:
        circuit_png = plot_dir / "00_equivalent_circuit.png"
        _draw_equivalent_circuit_image(p, circuit_png)
        cimg = XLImage(str(circuit_png))
        cimg.width = 980
        cimg.height = 560
        ws_param.add_image(cimg, "A20")

        zcorr_mag = np.abs(Zcorr)
        zcorr_phase = result["Zcorr_phase_tau_deg"]
        zmodel_mag = np.abs(Zmodel)
        zmodel_phase = np.rad2deg(np.angle(Zmodel))
        copen_pf = result["Copen_pF"]

        _add_plot_image_sheet(
            ws_amp,
            "SC-05-30J 周波数特性（振幅・フィットモデル重ね書き）",
            "補正後の測定|Z|と、今回フィットしなおした等価回路モデルの|Z|を比較。",
            plot_dir / "01_amp_full.png",
            f,
            [
                {"y": zcorr_mag, "label": "補正後 |Z|", "style": "meas"},
                {"y": zmodel_mag, "label": "今回フィットモデル |Z|", "style": "model"},
            ],
            "|Z| [Ω]",
            _format_ohm,
            log_y=True,
        )
        _add_plot_image_sheet(
            ws_phase,
            "SC-05-30J 周波数特性（位相・フィットモデル重ね書き）",
            "測定位相は補正後+τ補正。モデル位相はτ補正なしで表示。",
            plot_dir / "02_phase_full.png",
            f,
            [
                {"y": zcorr_phase, "label": "補正後+τ補正 位相", "style": "meas"},
                {"y": zmodel_phase, "label": "モデル位相 τ補正なし", "style": "model"},
            ],
            "Phase [deg]",
            _format_deg,
            log_y=False,
            y_min=-180,
            y_max=180,
        )
        _add_plot_image_sheet(
            ws_open,
            "Zopen測定値から求めたOpen等価容量",
            "Zopenのアドミタンスから Copen = Im(Yopen)/(2πf) として算出。",
            plot_dir / "03_open_capacity.png",
            f,
            [
                {"y": copen_pf, "label": "Copen", "style": "open"},
            ],
            "Copen [pF]",
            _format_pf,
            log_y=False,
        )
        _add_plot_image_sheet(
            ws_peak,
            "主自己共振拡大（今回モデル重ね書き）",
            "主自己共振付近を拡大。横軸は周波数、縦軸は|Z|。",
            plot_dir / "04_main_resonance_zoom.png",
            f,
            [
                {"y": zcorr_mag, "label": "補正後 |Z|", "style": "meas"},
                {"y": zmodel_mag, "label": "今回フィットモデル |Z|", "style": "model"},
            ],
            "|Z| [Ω]",
            _format_ohm,
            log_y=True,
            x_min=max(float(f.min()), fit.f_main_peak_hz / 3.0),
            x_max=min(float(f.max()), fit.f_main_peak_hz * 3.0),
        )
        hf_sheet_title = "高周波側共振拡大（今回モデル重ね書き）"
        ws_hf.title = "グラフ_20MHz拡大"
        _add_plot_image_sheet(
            ws_hf,
            hf_sheet_title,
            "10〜25MHz付近の高周波側共振/谷を拡大。横軸は周波数、縦軸は|Z|。",
            plot_dir / "05_hf_resonance_zoom.png",
            f,
            [
                {"y": zcorr_mag, "label": "補正後 |Z|", "style": "meas"},
                {"y": zmodel_mag, "label": "今回フィットモデル |Z|", "style": "model"},
            ],
            "|Z| [Ω]",
            _format_ohm,
            log_y=True,
            x_min=max(float(f.min()), 1.0e7),
            x_max=min(float(f.max()), 2.5e7),
        )

        wb.save(output_path)
    finally:
        # openpyxlはsave時に画像をxlsx内へ取り込むため、保存後に一時画像は削除可能。
        import shutil
        shutil.rmtree(plot_dir, ignore_errors=True)

    return output_path


def analyze_and_export(
    meas_csv: str | Path,
    open_csv: str | Path,
    td_ns: float,
    rref_ohm: float = RREF_OHM_DEFAULT,
    short_csv: str | Path | None = None,
    short_mode: str = "open_only",
    rs_mode: str = RS_MODE_DEFAULT,
    rs_min_ohm: float = RS_MIN_OHM_DEFAULT,
    rs_fixed_ohm: float = RS_FIXED_OHM_DEFAULT,
) -> Path:
    result = analyze(
        meas_csv,
        open_csv,
        td_ns,
        rref_ohm,
        short_csv=short_csv,
        short_mode=short_mode,
        rs_mode=rs_mode,
        rs_min_ohm=rs_min_ohm,
        rs_fixed_ohm=rs_fixed_ohm,
    )
    meas_path = Path(meas_csv)
    suffix = f"_rs_{result['rs_mode']}"
    out_name = f"{meas_path.stem}_open_tau_corrected_fit{suffix}.xlsx"
    out_path = meas_path.with_name(out_name)
    return export_excel(result, out_path)


# -----------------------------------------------------------------------------
# GUI
# -----------------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("SC-05-30J インピーダンス解析・フィッティング")
        self.root.geometry("900x650")

        self.meas_path = tk.StringVar()
        self.open_path = tk.StringVar()
        self.short_path = tk.StringVar()
        self.short_mode = tk.StringVar(value="open_only")
        self.rs_mode = tk.StringVar(value=RS_MODE_DEFAULT)
        self.rs_min_ohm = tk.StringVar(value=str(RS_MIN_OHM_DEFAULT))
        self.rs_fixed_ohm = tk.StringVar(value=str(RS_FIXED_OHM_DEFAULT))
        self.td_ns = tk.StringVar(value=str(TD_NS_DEFAULT))
        self.rref = tk.StringVar(value=str(RREF_OHM_DEFAULT))
        self.status = tk.StringVar(value="未実行")

        padx = 12
        pady = 8

        title = tk.Label(root, text="Open/Short補正・τ補正・等価回路フィット Excel出力 v13", font=("Meiryo UI", 14, "bold"))
        title.pack(anchor="w", padx=padx, pady=(14, 8))

        frame = tk.Frame(root)
        frame.pack(fill="x", padx=padx, pady=4)

        btn_meas = tk.Button(frame, text="測定データ採取", width=18, command=self.select_meas)
        btn_meas.grid(row=0, column=0, padx=(0, 8), pady=pady, sticky="w")
        ent_meas = tk.Entry(frame, textvariable=self.meas_path, width=82)
        ent_meas.grid(row=0, column=1, pady=pady, sticky="we")

        btn_open = tk.Button(frame, text="Zopenデータ採取", width=18, command=self.select_open)
        btn_open.grid(row=1, column=0, padx=(0, 8), pady=pady, sticky="w")
        ent_open = tk.Entry(frame, textvariable=self.open_path, width=82)
        ent_open.grid(row=1, column=1, pady=pady, sticky="we")

        btn_short = tk.Button(frame, text="Zshortデータ採取", width=18, command=self.select_short)
        btn_short.grid(row=2, column=0, padx=(0, 8), pady=pady, sticky="w")
        ent_short = tk.Entry(frame, textvariable=self.short_path, width=82)
        ent_short.grid(row=2, column=1, pady=pady, sticky="we")
        tk.Label(frame, text="未選択なら従来のOpen補正のみ", fg="#666666").grid(row=3, column=1, sticky="w")

        opt = tk.Frame(root)
        opt.pack(fill="x", padx=padx, pady=4)
        tk.Label(opt, text="位相補償 td [ns]").grid(row=0, column=0, padx=(0, 8), sticky="w")
        tk.Entry(opt, textvariable=self.td_ns, width=12).grid(row=0, column=1, padx=(0, 24), sticky="w")
        tk.Label(opt, text="Rref [Ω]").grid(row=0, column=2, padx=(0, 8), sticky="w")
        tk.Entry(opt, textvariable=self.rref, width=12).grid(row=0, column=3, padx=(0, 24), sticky="w")
        tk.Label(opt, text="通常は td=2.6 ns, Rref=49.8 Ω（7/1版Excelと同じ）").grid(row=0, column=4, sticky="w")

        mode_frame = tk.LabelFrame(root, text="Zshort補正モード", padx=8, pady=6)
        mode_frame.pack(fill="x", padx=padx, pady=(4, 4))
        tk.Radiobutton(
            mode_frame,
            text="Open補正のみ（Zshortは診断値としてExcel出力）",
            variable=self.short_mode,
            value="open_only",
        ).grid(row=0, column=0, sticky="w", padx=(0, 16))
        tk.Radiobutton(
            mode_frame,
            text="低周波一定R/L補正（推奨・過補正しにくい）",
            variable=self.short_mode,
            value="lowfreq_rl",
        ).grid(row=0, column=1, sticky="w", padx=(0, 16))
        tk.Radiobutton(
            mode_frame,
            text="全周波数Zshort補正（参考・20MHz付近で過補正注意）",
            variable=self.short_mode,
            value="full",
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(4, 0))

        rs_frame = tk.LabelFrame(root, text="Rsフィット制約", padx=8, pady=6)
        rs_frame.pack(fill="x", padx=padx, pady=(4, 4))
        tk.Radiobutton(
            rs_frame,
            text="自由フィット（従来互換）",
            variable=self.rs_mode,
            value="free",
        ).grid(row=0, column=0, sticky="w", padx=(0, 16))
        tk.Radiobutton(
            rs_frame,
            text="下限付きフィット",
            variable=self.rs_mode,
            value="lower_bound",
        ).grid(row=0, column=1, sticky="w", padx=(0, 8))
        tk.Label(rs_frame, text="Rs_min [Ω]").grid(row=0, column=2, sticky="e", padx=(8, 4))
        tk.Entry(rs_frame, textvariable=self.rs_min_ohm, width=10).grid(row=0, column=3, sticky="w", padx=(0, 16))
        tk.Radiobutton(
            rs_frame,
            text="固定値でフィット",
            variable=self.rs_mode,
            value="fixed",
        ).grid(row=1, column=0, sticky="w", pady=(4, 0), padx=(0, 16))
        tk.Label(rs_frame, text="Rs_fixed [Ω]").grid(row=1, column=1, sticky="e", pady=(4, 0), padx=(8, 4))
        tk.Entry(rs_frame, textvariable=self.rs_fixed_ohm, width=10).grid(row=1, column=2, sticky="w", pady=(4, 0), padx=(0, 16))
        tk.Label(rs_frame, text="推奨初期値: 下限付き 0.05 Ω。DC抵抗が分かる場合は固定値モード。", fg="#666666").grid(row=2, column=0, columnspan=4, sticky="w", pady=(4, 0))

        run_frame = tk.Frame(root)
        run_frame.pack(fill="x", padx=padx, pady=(8, 4))
        self.fit_button = tk.Button(run_frame, text="フィッティング", width=18, height=2, command=self.run_fit)
        self.fit_button.pack(side="left", padx=(0, 12))
        self.status_label = tk.Label(run_frame, textvariable=self.status, anchor="w", fg="gray", font=("Meiryo UI", 10, "bold"))
        self.status_label.pack(side="left", fill="x", expand=True)

        note = (
            "処理順: CSV複素化 → Z = Rref·H/(1-H) → 選択モードに応じたZshort補正 "
            "→ Openアドミタンス補正 → 測定値だけτ補正 → 等価回路フィット → Excel保存\n"
            "デフォルトはOpen補正のみです。Zshortは診断値としてExcelに出力されます。モデル位相にはτ補正を加えません。\n"
            "Rsはデフォルトで下限付きフィットです。全周波数Zshort補正でRsが非物理的に小さくなる場合は、Rs_minまたは固定値を設定してください。"
        )
        tk.Label(root, text=note, justify="left", fg="#444444", wraplength=740).pack(anchor="w", padx=padx, pady=(10, 0))

    def select_meas(self):
        path = filedialog.askopenfilename(
            title="測定データCSVを選択",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            self.meas_path.set(path)

    def select_open(self):
        initial_dir = str(Path(self.meas_path.get()).parent) if self.meas_path.get() else None
        path = filedialog.askopenfilename(
            title="ZopenデータCSVを選択",
            initialdir=initial_dir,
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            self.open_path.set(path)

    def select_short(self):
        initial_dir = str(Path(self.meas_path.get()).parent) if self.meas_path.get() else None
        path = filedialog.askopenfilename(
            title="ZshortデータCSVを選択（任意）",
            initialdir=initial_dir,
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            self.short_path.set(path)

    def run_fit(self):
        try:
            meas = self.meas_path.get().strip()
            opn = self.open_path.get().strip()
            sht = self.short_path.get().strip()
            if not meas or not Path(meas).exists():
                messagebox.showerror("エラー", "測定データCSVを選択してください。")
                return
            if not opn or not Path(opn).exists():
                messagebox.showerror("エラー", "ZopenデータCSVを選択してください。")
                return
            if sht and not Path(sht).exists():
                messagebox.showerror("エラー", "ZshortデータCSVが見つかりません。未使用なら入力欄を空にしてください。")
                return
            short_csv = sht if sht else None
            short_mode = self.short_mode.get()
            rs_mode = self.rs_mode.get()
            rs_min_ohm = float(self.rs_min_ohm.get())
            rs_fixed_ohm = float(self.rs_fixed_ohm.get())
            td_ns = float(self.td_ns.get())
            rref = float(self.rref.get())

            self.status.set("フィッティング中...")
            self.status_label.config(fg="blue")
            self.fit_button.config(state="disabled")
            self.root.update_idletasks()

            out = analyze_and_export(
                meas,
                opn,
                td_ns,
                rref,
                short_csv=short_csv,
                short_mode=short_mode,
                rs_mode=rs_mode,
                rs_min_ohm=rs_min_ohm,
                rs_fixed_ohm=rs_fixed_ohm,
            )

            self.status.set(f"完了: {out}")
            self.status_label.config(fg="green")
            messagebox.showinfo("完了", f"Excelを出力しました。\n\n{out}")
        except Exception as exc:
            self.status.set("エラー")
            self.status_label.config(fg="red")
            msg = f"処理中にエラーが発生しました。\n\n{exc}\n\n詳細:\n{traceback.format_exc()}"
            messagebox.showerror("エラー", msg)
        finally:
            self.fit_button.config(state="normal")


def main() -> None:
    if tk is None:
        raise RuntimeError(f"tkinterを読み込めません: {_TK_IMPORT_ERROR}")
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()

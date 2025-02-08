import openpyxl
import csv
import time

def read_excel_lxml(excel_file, sheet_name, csv_file):
    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name]

    # CSVファイルに書き出す
    with open(csv_file, mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)

        # ヘッダーを設定 (作業日、氏名、工事番号、実績工数、所属)
        writer.writerow(["作業日", "氏名", "工事番号", "実績工数", "所属"])

        # 2行目からデータを取得
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # 行ごとのデータを取得 (必要な列だけ)
            作業日 = row[0].value  # A列
            氏名 = row[1].value    # B列
            工事番号 = row[2].value  # C列
            実績工数 = row[3].value  # D列
            所属 = row[4].value   # E列

            # 行データとしてリストに格納
            row_data = [作業日, 氏名, 工事番号, 実績工数, 所属]

            # CSVファイルに行データを書き込む
            writer.writerow(row_data)

    print(f"CSVファイル '{csv_file}' にデータを書き出しました。")

# 使用例
excel_file = input("エクセルファイル名を入力してください: ")
sheet_name = input("シート名を入力してください: ")
csv_file = excel_file.replace('.xlsx', '.csv')

time_start = time.time()
print("<処理開始>")

read_excel_lxml(excel_file, sheet_name, csv_file)

time_end = time.time() - time_start
print("<処理時間> ", time_end)
